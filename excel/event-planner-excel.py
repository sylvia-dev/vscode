import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def create_event_planning_matrix(output_file="Event_Planning_Matrix.xlsx"):
    # Create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    # Create sample data for the main events sheet
    events_data = {
        'Event Name': ['Industry Conference A', 'Trade Show B', 'Partner Event C'],
        'Quarter': ['Q1', 'Q2', 'Q1'],
        'Date': ['2024-02-15', '2024-05-20', '2024-03-10'],
        'Location': ['New York', 'London', 'Singapore'],
        'Vertical': ['Finance', 'Technology', 'Healthcare'],
        'Country': ['USA', 'UK', 'Singapore'],
        'Event Type': ['Conference', 'Trade Show', 'Partner Event'],
        'Budget': [50000, 75000, 30000],
        'Expected Attendees': [500, 1000, 300],
        'Status': ['Confirmed', 'Planning', 'Tentative'],
        'Key Stakeholders': ['Sales Team', 'Product Team', 'Partnership Team'],
        'Notes': ['Keynote speaking slot', 'Booth size 20x20', 'Joint presentation']
    }

    # Create main events sheet
    df_events = pd.DataFrame(events_data)
    df_events.to_excel(writer, sheet_name='Events Master List', index=False)

    # Create quarterly summary sheet
    quarterly_summary = df_events.groupby('Quarter').agg({
        'Event Name': 'count',
        'Budget': 'sum',
        'Expected Attendees': 'sum'
    }).reset_index()
    quarterly_summary.columns = ['Quarter', 'Number of Events', 'Total Budget', 'Total Expected Attendees']
    quarterly_summary.to_excel(writer, sheet_name='Quarterly Summary', index=False)

    # Create vertical analysis sheet
    vertical_analysis = df_events.groupby('Vertical').agg({
        'Event Name': 'count',
        'Budget': 'sum',
        'Expected Attendees': 'sum'
    }).reset_index()
    vertical_analysis.columns = ['Vertical', 'Number of Events', 'Total Budget', 'Total Expected Attendees']
    vertical_analysis.to_excel(writer, sheet_name='Vertical Analysis', index=False)

    # Create geographic analysis sheet
    geo_analysis = df_events.groupby('Country').agg({
        'Event Name': 'count',
        'Budget': 'sum',
        'Expected Attendees': 'sum'
    }).reset_index()
    geo_analysis.columns = ['Country', 'Number of Events', 'Total Budget', 'Total Expected Attendees']
    geo_analysis.to_excel(writer, sheet_name='Geographic Analysis', index=False)

    # Access the workbook and the sheets
    workbook = writer.book
    
    # Format all sheets
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        # Format header row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    # Add color scaling to budget columns
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        budget_col = None
        
        # Find the budget column
        for idx, cell in enumerate(ws[1], 1):
            if 'Budget' in str(cell.value):
                budget_col = get_column_letter(idx)
                break
        
        if budget_col:
            # Add color scale rule
            color_scale_rule = ColorScaleRule(
                start_type='min',
                start_color='63BE7B',  # Green
                mid_type='percentile',
                mid_value=50,
                mid_color='FFEB84',  # Yellow
                end_type='max',
                end_color='F8696B'   # Red
            )
            ws.conditional_formatting.add(f'{budget_col}2:{budget_col}{ws.max_row}', color_scale_rule)

    # Save the workbook
    writer.close()

    return output_file

# Generate the Excel file
output_file = create_event_planning_matrix()
print(f"Excel file created: {output_file}")
